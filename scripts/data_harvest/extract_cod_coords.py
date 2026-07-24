"""Extract BBS-coded centroids from the OCHA COD-AB Bangladesh gazetteer.

Reads bgd_admin.xlsx (data.humdata.org cod-ab-bgd), writes cod_coords.json:
{
  "districts": {"5081": [lat, lon], ...},          # admin2 center_lat/lon
  "upazilas":  {"508194": [lat, lon], ...},        # adminpoints level 3
  "unions":    {"50819427": [lat, lon], ...}       # adminpoints level 4
}
Pcodes are "BD" + BBS geocode -> stripped to the bare geocode used by CZIS.
"""
import json

import openpyxl

wb = openpyxl.load_workbook("bgd_admin.xlsx", read_only=True)

out = {"districts": {}, "upazilas": {}, "unions": {}}

ws = wb["bgd_admin2"]
rows = list(ws.iter_rows(values_only=True))
i = {h: n for n, h in enumerate(rows[0])}
for r in rows[1:]:
    pcode = str(r[i["adm2_pcode"]] or "")
    if pcode.startswith("BD") and r[i["center_lat"]] is not None:
        out["districts"][pcode[2:]] = [
            round(float(r[i["center_lat"]]), 5),
            round(float(r[i["center_lon"]]), 5),
        ]

ws = wb["bgd_adminpoints"]
rows = list(ws.iter_rows(values_only=True))
i = {h: n for n, h in enumerate(rows[0])}
for r in rows[1:]:
    level = r[i["admin_level"]]
    x, y = r[i["x_coord"]], r[i["y_coord"]]
    if x is None or y is None:
        continue
    lat, lon = round(float(y), 5), round(float(x), 5)
    if level == 3:
        pcode = str(r[i["adm3_pcode"]] or "")
        if pcode.startswith("BD"):
            out["upazilas"][pcode[2:]] = [lat, lon]
    elif level == 4:
        pcode = str(r[i["adm4_pcode"]] or "")
        if pcode.startswith("BD"):
            out["unions"][pcode[2:]] = [lat, lon]

with open("cod_coords.json", "w") as f:
    json.dump(out, f)

print(
    f"districts={len(out['districts'])} upazilas={len(out['upazilas'])} "
    f"unions={len(out['unions'])}"
)
print("Tanore 508194:", out["upazilas"].get("508194"))
print("Badhair 50819427:", out["unions"].get("50819427"))
